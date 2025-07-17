from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook("regional_sales_summary.xlsx")
ws = wb.active

for cell in ws[1]:
    cell.font = Font(bold=True)

wb.save("regional_sales_summary.xlsx")

print("Header formatting applied.")